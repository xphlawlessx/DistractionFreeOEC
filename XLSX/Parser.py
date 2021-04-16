from openpyxl import load_workbook
from openpyxl.writer.excel import save_workbook
from openpyxl.utils import column_index_from_string


class Code():
    def __init__(self, code_name_, cell_, color_):
        self.code_name = code_name_
        self.cell = cell_
        self.color = color_

    def __str__(self):
        return f'{self.code_name} - {self.cell} - {self.color}'

    def __repr__(self):
        return f'{self.code_name} - {self.cell} - {self.color}'


class XLSXParser:

    def __init__(self):
        self.work_book = None
        self.sheet = ''
        self.question = ''
        self.question_cell = ''
        self.first_comment_col = ''
        self.first_comment_index = 5
        self.comment_index = 5
        self.first_code_col = ''
        self.old_data = []

    def get_sheet_names(self) -> list:
        if self.work_book is None:
            []
        return self.work_book.sheetnames

    def load_workbook(self, filename):
        self.work_book = load_workbook(filename, data_only=True)

    def get_completed_code_values(self, completed_codes):

        # vals = [''.join(str(p[1]).split('-')[1:]).lstrip() for c in completed_codes.values() for p in c]
        vals = []
        for c in completed_codes.values():
            for p in c:
                if str(p[1]).count('-') > 1:
                    vals.append('-'.join(str(p[1]).split('-')[1:]).lstrip())
                else:
                    vals.append(''.join(str(p[1]).split('-')[1:]).lstrip())

        return vals

    def save_workbook(self, filename: str, completed_codes: dict):
        self.comment_index = self.first_comment_index
        cell = self.first_comment_col + str(self.comment_index)
        val = self.sheet[cell].value
        vals = self.get_completed_code_values(completed_codes)
        comment_count = len(vals)
        change_count = 0

        def get_cols() -> dict:
            col_count = 0
            cols = {}
            for col in self.sheet.iter_cols(min_row=3, max_row=3):
                for cell in col:
                    if cell.value in vals:
                        cols[cell.value] = ''.join(list(str(cell).split('.')[-1])[:-2])
                        col_count += 1
                        if col_count == comment_count:
                            return cols

        cols = get_cols()
        while change_count < comment_count:
            if val in completed_codes.keys():
                for key in vals:
                    _cell = cols[key] + str(self.comment_index)
                    self.sheet[_cell].value = 1.0
                    change_count += 1
            self.next_comment()
            cell = self.first_comment_col + str(self.comment_index)
            val = self.sheet[cell].value
        save_workbook(self.work_book, filename)

    def get_codes(self, sheet_name):
        self.sheet = self.work_book[sheet_name]
        codes = []
        for col in self.sheet.iter_cols(min_row=3, max_row=3, min_col=8):
            for cell in col:
                if cell.value:
                    color_str = str(cell.fill).replace('<openpyxl.styles.fills.PatternFill object>', '').replace('\n',
                                                                                                                 '')
                    check_str = str(cell.value).lower()
                    if check_str.find('net:') < 0 and check_str != 'count' and check_str.find(sheet_name.lower()) < 0:
                        if self.first_code_col == '':
                            self.first_code_col = column_index_from_string(
                                str(col).split('.')[-1].replace('>,)', '')[0])

                        codes.append(
                            Code(cell.value, column_index_from_string(str(col).split('.')[-1].replace('>,)', '')[0]),
                                 color_str))
        self.set_question(sheet_name)
        return codes

    def parse_old_data(self):
        for row in self.sheet.iter_rows(min_col=self.first_code_col, min_row=self.first_comment_index):
            for cell in row:
                if cell.value:
                    val = str(cell).split('.')[-1]
                    self.old_data.append((column_index_from_string(val[0]) - self.first_code_col,
                                          self.sheet[self.first_comment_col + val[1]].value,
                                          int(val[1])))
                    #  [(1, '0 - Environment - general / vague')

    def set_question(self, sheet_name):
        for col in self.sheet.iter_cols(max_col=10, max_row=10):
            for cell in col:
                if cell.value:
                    if str(cell.value).lower().find(sheet_name.lower()) != -1:
                        self.question = cell.value
                        self.question_cell = str(cell).split('.')[-1].replace('>', '')
                        splits = list(self.question_cell)
                        self.first_comment_col = ''.join(splits[:-1])
                        self.comment_index = self.first_comment_index

    def get_current_comment(self):
        cell = self.first_comment_col + str(self.comment_index)
        return self.sheet[cell].value

    def next_comment(self):
        self.comment_index += 1

    def prev_comment(self):
        self.comment_index -= 1

    @staticmethod
    def clamp(n, smallest, largest):
        return max(smallest, min(n, largest))
